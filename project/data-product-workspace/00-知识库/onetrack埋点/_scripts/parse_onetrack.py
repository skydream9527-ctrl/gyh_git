#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""解析 OneTrack 埋点 Excel,输出结构化 JSON。
- 自动识别表头行(含"事件名"的行)
- 合并单元格:事件名向下填充
- 识别公共属性引用行(common key / content key 等)
- 提取字体颜色作为版本标记(文件1: 黑=一期/紫=二期/红=二期重点/蓝=三期重点/黄底=新增)
"""
import json, sys, re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 公共属性引用行的标识(属性名英文列出现这些值时,表示引用公共属性,非独立参数)
PRESET_REF_TOKENS = {'common key', 'commen key', 'commey_key', 'comment_key', 'content key', 'content_key'}

# 列名匹配规则(关键字 -> 标准字段)
def match_col(header_val):
    if header_val is None:
        return None
    s = str(header_val).strip().lower().replace(' ', '').replace('\n', '').replace('\r', '').replace('（', '(').replace('）', ')')
    if '事件名(英文)' in s or s == 'eventname(english)' or 'eventname(english)' in s:
        return 'event_en'
    if '事件名(中文)' in s or 'eventname(chinese)' in s:
        return 'event_cn'
    if '上报时机' in s or '上报逻辑' in s or 'reporttime' in s or 'reportlogic' in s:
        return 'report'
    if '属性名(英文)' in s or 'keyname(english)' in s:
        return 'param_en'
    if '属性名(中文)' in s or 'keyname(chinese)' in s:
        return 'param_cn'
    if '属性值类型' in s or 'keyvaluetype' in s:
        return 'value_type'
    if '属性名的对应值' in s or 'keyvalue&othernote' in s or 'keyvalue&other' in s:
        return 'value_desc'
    if '进版版本号' in s or '带出版本' in s or 'app_ver' in s.replace('_','') or s == '版本':
        return 'version'
    if '无痕模式是否上报' in s:
        return 'incognito'
    if '备注' in s and 'remark' in s:
        return 'remark'
    if s == '分类' or s == 'classification' or s == 'type' or s == 'keytype':
        return 'category'
    return None

def color_tag(cell):
    """提取字体颜色/填充色作为版本标记"""
    tags = []
    try:
        f = cell.font
        if f and f.color and f.color.rgb:
            rgb = str(f.color.rgb)
            # ARGB hex
            if rgb and rgb != '00000000' and rgb != 'FF000000':
                # 简单判定主色
                r = int(rgb[-6:-4], 16) if len(rgb) >= 6 else 0
                g = int(rgb[-4:-2], 16) if len(rgb) >= 6 else 0
                b = int(rgb[-2:], 16) if len(rgb) >= 6 else 0
                if b > 150 and r < 100 and g < 100:
                    tags.append('blue(三期重点)')
                elif r > 150 and g < 80 and b > 80:
                    tags.append('purple(二期)')
                elif r > 180 and g < 80 and b < 80:
                    tags.append('red(二期重点)')
        patt = cell.fill
        if patt and patt.fgColor and patt.fgColor.rgb:
            rgb = str(patt.fgColor.rgb)
            if rgb and rgb != '00000000' and rgb != 'FFFFFFFF' and rgb != '00FFFFFF':
                if len(rgb) >= 6:
                    r = int(rgb[-6:-4], 16)
                    g = int(rgb[-4:-2], 16)
                    b = int(rgb[-2:], 16)
                    if r > 230 and g > 200 and b < 100:
                        tags.append('yellow_bg(新增字段)')
    except Exception:
        pass
    return tags

def clean(v):
    if v is None:
        return ''
    s = str(v).replace('​', '').replace('\xa0', ' ').strip()
    if s.lower() == 'nan':
        return ''
    return s

def looks_like_header(ev_en, p_en, v_type, report):
    """判断是否仍是表头行(中英文表头残留)"""
    e = ev_en.lower()
    p = p_en.lower()
    v = v_type.lower()
    r = report.lower()
    if 'event name' in e and ('english' in e or 'chinese' in e):
        return True
    if 'key name' in p and ('english' in p or 'chinese' in p):
        return True
    if 'key value type' in v:
        return True
    if 'key value & other note' in v or 'key value & other' in v:
        return True
    if 'report time' in r and 'report time' in e:
        return True
    if e == 'event name（english）' or e == 'event name（chinese）':
        return True
    # 中文表头残留
    CN_HEADERS = {'事件名（英文）', '事件名（中文）', '属性名（英文）', '属性名（中文）',
                  '上报时机', '上报逻辑', '属性值类型', '属性名的对应值及说明',
                  '备注', '进版版本号', '无痕模式是否上报', '分类', 'key type'}
    if ev_en in CN_HEADERS or p_en in CN_HEADERS:
        return True
    return False


def parse_sheet(ws):
    """解析单个 sheet,返回 {meta, events, is_preset_sheet, preset_rows}"""
    max_row, max_col = ws.max_row, ws.max_column
    # 读取所有单元格
    grid = {}
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for c in row:
            grid[(c.row, c.column)] = c

    # 保留原始文本行(前 80 行,便于说明性 sheet 可追溯)
    raw_lines = []
    for r in range(1, min(max_row + 1, 81)):
        parts = []
        for col in range(1, max_col + 1):
            if (r, col) in grid:
                v = clean(grid[(r, col)].value)
                if v:
                    parts.append(v)
        if parts:
            raw_lines.append(' | '.join(parts))

    # 找表头行:前 6 行中包含"事件名"关键字的行
    header_row = None
    col_map = {}
    preset_header = None
    for r in range(1, min(7, max_row + 1)):
        row_vals = {}
        for col in range(1, max_col + 1):
            if (r, col) in grid:
                std = match_col(grid[(r, col)].value)
                if std:
                    row_vals[std] = col
        # 优先:有 event_en 和 param_en 的事件表头
        if 'event_en' in row_vals and 'param_en' in row_vals:
            header_row = r
            col_map = row_vals
            break
        # 备选:有 param_en 无 event_en 的预置表头(不 break,继续找事件表头)
        if 'param_en' in row_vals and 'event_en' not in row_vals and len(row_vals) >= 2:
            if preset_header is None:
                preset_header = (r, row_vals)
    if header_row is None and preset_header is not None:
        header_row, col_map = preset_header

    if header_row is None:
        # 说明性 sheet,提取纯文本
        lines = []
        for r in range(1, min(max_row + 1, 60)):
            row_text = ' | '.join(clean(grid[(r, col)].value) for col in range(1, max_col + 1) if (r, col) in grid and clean(grid[(r, col)].value))
            if row_text.strip():
                lines.append(row_text)
        return {'type': 'note', 'header_row': None, 'lines': lines, 'raw_lines': raw_lines, 'events': [], 'preset': []}

    has_event = 'event_en' in col_map
    events = []
    preset = []
    cur_event = None
    cur_category = ''

    for r in range(header_row + 1, max_row + 1):
        # 读分类(向下填充)
        if 'category' in col_map:
            cat_val = clean(grid[(r, col_map['category'])].value) if (r, col_map['category']) in grid else ''
            if cat_val:
                cur_category = cat_val

        ev_en_col = col_map.get('event_en')
        ev_cn_col = col_map.get('event_cn')
        ev_en = clean(grid[(r, ev_en_col)].value) if has_event and ev_en_col and (r, ev_en_col) in grid else ''
        ev_cn = clean(grid[(r, ev_cn_col)].value) if has_event and ev_cn_col and (r, ev_cn_col) in grid else ''
        # 事件名可能多个换行(多个事件共用参数),取第一个为主,其余记为别名
        ev_en_first = ev_en.split('\n')[0].strip() if ev_en else ''
        ev_cn_first = ev_cn.split('\n')[0].strip() if ev_cn else ''

        p_en_col = col_map.get('param_en')
        p_cn_col = col_map.get('param_cn')
        v_type_col = col_map.get('value_type')
        v_desc_col = col_map.get('value_desc')
        report_col = col_map.get('report')
        remark_col = col_map.get('remark')
        version_col = col_map.get('version')
        incognito_col = col_map.get('incognito')
        p_en = clean(grid[(r, p_en_col)].value) if p_en_col and (r, p_en_col) in grid else ''
        p_cn = clean(grid[(r, p_cn_col)].value) if p_cn_col and (r, p_cn_col) in grid else ''
        v_type = clean(grid[(r, v_type_col)].value) if v_type_col and (r, v_type_col) in grid else ''
        v_desc = clean(grid[(r, v_desc_col)].value) if v_desc_col and (r, v_desc_col) in grid else ''
        report = clean(grid[(r, report_col)].value) if report_col and (r, report_col) in grid else ''
        remark = clean(grid[(r, remark_col)].value) if remark_col and (r, remark_col) in grid else ''
        version = clean(grid[(r, version_col)].value) if version_col and (r, version_col) in grid else ''
        incognito = clean(grid[(r, incognito_col)].value) if incognito_col and (r, incognito_col) in grid else ''

        # 空行跳过
        if not any([ev_en, p_en, p_cn, v_type, v_desc, report, remark, version]):
            continue

        # 跳过表头残留行(中英文表头)
        if has_event and looks_like_header(ev_en, p_en, v_type, report):
            continue
        if not has_event and ('key name' in p_en.lower() and ('english' in p_en.lower() or 'chinese' in p_en.lower())):
            continue
        if not has_event and 'key value type' in v_type.lower():
            continue

        if not has_event:
            # 预置属性 sheet
            if p_en or p_cn:
                preset.append({
                    'param_en': p_en,
                    'param_cn': p_cn,
                    'value_type': v_type,
                    'value_desc': v_desc,
                    'remark': remark,
                    'version': version,
                })
            continue

        # 事件表
        if ev_en_first:
            cur_event = {
                'event_en': ev_en_first,
                'event_cn': ev_cn_first,
                'event_en_raw': ev_en,
                'event_cn_raw': ev_cn,
                'report': report,
                'remark': remark,
                'version': version,
                'incognito': incognito,
                'category': cur_category,
                'params': [],
                'preset_refs': [],
            }
            events.append(cur_event)
        if cur_event is None:
            continue

        # 处理参数行
        p_en_low = p_en.lower().replace(' ', '').replace('_', '')
        if p_en_low in PRESET_REF_TOKENS or p_en in PRESET_REF_TOKENS or p_en_low in {t.replace(' ','').replace('_','') for t in PRESET_REF_TOKENS}:
            ref = p_en if p_en else p_cn
            if ref and ref not in cur_event['preset_refs']:
                cur_event['preset_refs'].append(ref)
            continue

        if p_en or p_cn or v_type or v_desc:
            # 字体颜色标记(仅文件1有意义)
            c_tags = []
            if p_en_col and (r, p_en_col) in grid:
                c_tags = color_tag(grid[(r, p_en_col)])
            cur_event['params'].append({
                'param_en': p_en,
                'param_cn': p_cn,
                'value_type': v_type,
                'value_desc': v_desc,
                'remark': remark,
                'version': version,
                'color_tag': c_tags,
            })

    # 后置降级:若 event_en 列实为类型分类(如"预置参数""公共参数"),转为 preset
    if has_event and events:
        uniq_ev = set(e['event_en'] for e in events if e['event_en'])
        if len(uniq_ev) <= 3:
            def looks_like_type(ev):
                if not ev:
                    return True
                # 含下划线(如 app_open)或纯小写英文 -> 像事件名
                if '_' in ev and re.match(r'^[a-z][a-z0-9_]*$', ev):
                    return False
                # 中文字符或以"参数/属性"结尾 -> 像类型分类
                if any('一' <= ch <= '鿿' for ch in ev):
                    return True
                if ev.endswith('参数') or ev.endswith('属性'):
                    return True
                return False
            if all(looks_like_type(ev) for ev in uniq_ev):
                # 降级:event_en 作为 category,参数合并到 preset
                for e in events:
                    cat = e['event_en'] or e.get('category', '')
                    for p in e['params']:
                        p['category'] = cat
                        preset.append(p)
                events = []
                has_event = False

    return {
        'type': 'preset' if not has_event else 'events',
        'header_row': header_row,
        'col_map': col_map,
        'events': events,
        'preset': preset,
        'raw_lines': raw_lines,
    }


def parse_file(path):
    wb = load_workbook(path, data_only=True)
    out = {'file': path.split('/')[-1], 'sheets': []}
    for sn in wb.sheetnames:
        ws = wb[sn]
        res = parse_sheet(ws)
        res['sheet'] = sn
        out['sheets'].append(res)
    wb.close()
    return out


if __name__ == '__main__':
    files = [
        '/Users/mi/Downloads/内容中心onetrack埋点(三期)  带沉浸式.xlsx',
        '/Users/mi/Downloads/浏览器新版OneTrack埋点汇总（可用于神策、数鲸）.xlsx',
    ]
    for f in files:
        data = parse_file(f)
        tag = 'content_center' if '内容中心' in f else 'browser'
        out_path = '/tmp/onetrack_%s.json' % tag
        with open(out_path, 'w', encoding='utf-8') as fp:
            json.dump(data, fp, ensure_ascii=False, indent=2)
        # 统计
        n_events = 0
        n_params = 0
        n_preset = 0
        for s in data['sheets']:
            n_events += len(s['events'])
            for e in s['events']:
                n_params += len(e['params'])
            n_preset += len(s['preset'])
        print('[%s] sheets=%d events=%d params=%d preset=%d -> %s' % (tag, len(data['sheets']), n_events, n_params, n_preset, out_path))
